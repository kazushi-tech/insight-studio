#!/usr/bin/env python3
"""
Excel KPI Debug Tool for ads-insights

Diagnoses Excel file structure and KPI column mapping status.

Usage:
    python debug_excel.py <path-to-excel-file>

Example:
    python debug_excel.py data/users/test/client/report.xlsx
"""

import sys
from pathlib import Path

# Add project root to path for imports
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "web" / "app"))

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# KPI_SPECS from kpi_extractor.py (keep in sync)
KPI_SPECS = [
    ("cost", ["費用", "広告費", "ご利用金額", "コスト", "Cost", "Spend", "ご利用額", "利用額", "使用金額", "消費額", "広告支出", "金額"]),
    ("impr", ["表示回数", "インプレッション", "Imp", "Impressions", "Impr", "表示", "インプレッション数", "イムプレッション", "imp数"]),
    ("click", ["クリック", "クリック数", "Clicks", "Click", "クリック回数"]),
    ("cv", ["CV", "コンバージョン", "Conversions", "獲得", "成約", "獲得件数", "獲得数", "購入数", "成約数", "到達", "申込", "予約", "申し込み件数", "購入完了件数", "完了数", "件数"]),
    ("conversion_value", ["CV値", "コンバージョン値", "Conversion value", "Conv. value", "Value", "合計コンバージョン値", "Total conversion value", "コンバージョンの価値", "価値", "値", "conv値", "CV価値", "売却額"]),
    ("revenue", ["売上", "売上高", "Revenue", "Sales", "収益", "総売上", "Total revenue", "購入額", "購入金額", "売上金額", "収益金額", "売却金額"]),
    ("ctr", ["CTR", "クリック率", "Click through rate"]),
    ("cvr", ["CVR", "コンバージョン率", "獲得率", "Conversion rate", "転換率"]),
    ("cpa", ["CPA", "獲得単価", "Cost / conv.", "Cost/conv.", "コンバージョン単価", "顧客獲得単価"]),
    ("cpc", ["CPC", "クリック単価", "Cost per click"]),
    ("roas", ["ROAS", "広告費用対効果", "投資対効果", "Return on Ad Spend"]),
    ("revenue_per_cv", ["売上単価", "1件あたり売上", "売上/CV"]),
]


def find_header_row(sheet, max_rows=50):
    """Find the header row by looking for KPI keywords."""
    all_synonyms = set()
    for _, synonyms in KPI_SPECS:
        all_synonyms.update(s.lower() for s in synonyms)

    for row_idx in range(1, min(max_rows + 1, sheet.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(50, sheet.max_column + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                row_values.append(str(cell.value).strip().lower())

        # Check if this row has at least 2 KPI-related headers
        matches = sum(1 for v in row_values if v in all_synonyms)
        if matches >= 2:
            return row_idx

    return None


def map_columns(sheet, header_row):
    """Map columns to KPI types."""
    mapping = {}
    unmatched = []

    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        if not cell.value:
            continue

        col_name = str(cell.value).strip()
        col_name_lower = col_name.lower()

        matched = False
        for kpi_key, synonyms in KPI_SPECS:
            if col_name_lower in [s.lower() for s in synonyms]:
                mapping[kpi_key] = (col_idx, col_name)
                matched = True
                break

        if not matched and col_name:
            unmatched.append((col_idx, col_name))

    return mapping, unmatched


def analyze_excel(filepath):
    """Analyze Excel file structure and KPI mapping."""
    print(f"\n{'='*60}")
    print(f"Analyzing: {filepath}")
    print(f"{'='*60}\n")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"Error loading file: {e}")
        return

    print(f"Sheets found: {len(wb.sheetnames)}")
    print("-" * 40)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n## Sheet: {sheet_name}")
        print(f"   Dimensions: {sheet.max_row} rows x {sheet.max_column} cols")

        header_row = find_header_row(sheet)
        if not header_row:
            print("   [!] No header row detected (no KPI keywords found)")
            continue

        print(f"   Header row: {header_row}")

        mapping, unmatched = map_columns(sheet, header_row)

        print("\n   KPI Column Mapping:")
        for kpi_key, synonyms in KPI_SPECS:
            if kpi_key in mapping:
                col_idx, col_name = mapping[kpi_key]
                print(f"   [OK] {kpi_key:15} -> Col {col_idx}: '{col_name}'")
            else:
                print(f"   [--] {kpi_key:15} -> NOT FOUND")

        if unmatched:
            print("\n   Unmatched Columns (potential new synonyms):")
            for col_idx, col_name in unmatched[:15]:  # Limit output
                print(f"   [?]  Col {col_idx}: '{col_name}'")
            if len(unmatched) > 15:
                print(f"   ... and {len(unmatched) - 15} more")

    print(f"\n{'='*60}\n")


def main():
    if len(sys.argv) < 2:
        print("Usage: python debug_excel.py <path-to-excel-file>")
        print("\nExample:")
        print("  python debug_excel.py data/users/test/client/report.xlsx")
        sys.exit(1)

    filepath = Path(sys.argv[1])
    if not filepath.exists():
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    if filepath.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
        print(f"Warning: File may not be an Excel file: {filepath.suffix}")

    analyze_excel(filepath)


if __name__ == "__main__":
    main()
