# image_extractor.py
# Excelファイルから画像（バナー）を抽出するモジュール

from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
import io
from PIL import Image
import base64


def extract_images_from_excel(file_path: Path, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Excelファイルから画像を抽出

    Args:
        file_path: Excelファイルパス
        sheet_name: シート名（指定しない場合は全シート）

    Returns:
        [
            {
                "sheet": シート名,
                "image_index": 画像インデックス,
                "anchor": セル位置（例: "A1"）,
                "width": 幅,
                "height": 高さ,
                "data": base64エンコードされた画像データ,
                "format": 画像形式（png/jpeg等）,
                "media_type": 媒体種別（fb/ydn/gdn/youtube等、推定）
            },
            ...
        ]
    """
    from openpyxl.utils import get_column_letter
    images = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        sheets_to_process = [wb[sheet_name]] if sheet_name else wb.worksheets

        print(f"[画像抽出] ファイル: {file_path.name}")
        print(f"[画像抽出] 全シート数: {len(wb.worksheets)}")
        print(f"[画像抽出] 処理対象: {len(sheets_to_process)}シート")

        for ws in sheets_to_process:
            print(f"[画像抽出] シート: {ws.title}")

            # 媒体種別を推定
            media_type = "unknown"
            sheet_lower = ws.title.lower()
            if "fb" in sheet_lower or "facebook" in sheet_lower:
                media_type = "Facebook"
            elif "ydn" in sheet_lower or "yahoo" in sheet_lower:
                media_type = "YDN"
            elif "gdn" in sheet_lower or "google" in sheet_lower:
                media_type = "GDN"
            elif "youtube" in sheet_lower or "yt" in sheet_lower:
                media_type = "YouTube"
            elif "デマンド" in ws.title or "demand" in sheet_lower:
                media_type = "デマンド"
            elif "imgad" in sheet_lower:
                media_type = "画像広告"

            if not hasattr(ws, '_images') or not ws._images:
                print(f"  → 画像なし（_imagesが空）")
                continue

            print(f"  → 画像数: {len(ws._images)} 枚")

            for idx, img in enumerate(ws._images):
                try:
                    # 画像データを取得
                    image_data = img._data()

                    # PILで画像を読み込み
                    pil_img = Image.open(io.BytesIO(image_data))

                    # Base64エンコード
                    buffered = io.BytesIO()
                    img_format = pil_img.format or 'PNG'
                    pil_img.save(buffered, format=img_format)
                    img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')

                    # アンカー位置を取得（正しい計算）
                    anchor_cell = "不明"
                    try:
                        if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                            anchor_from = img.anchor._from
                            if hasattr(anchor_from, 'col') and hasattr(anchor_from, 'row'):
                                col_letter = get_column_letter(anchor_from.col + 1)  # 0-indexed
                                row_num = anchor_from.row + 1
                                anchor_cell = f"{col_letter}{row_num}"
                    except Exception as anchor_err:
                        print(f"  → アンカー取得エラー: {anchor_err}")

                    print(f"  → 画像{idx + 1}: {img_format}, {pil_img.width}x{pil_img.height}px, {anchor_cell}")

                    images.append({
                        "sheet": ws.title,
                        "image_index": idx,
                        "anchor": anchor_cell,
                        "width": pil_img.width,
                        "height": pil_img.height,
                        "data": img_base64,
                        "format": img_format.lower(),
                        "media_type": media_type,
                    })
                except Exception as e:
                    print(f"  → 画像抽出エラー (sheet={ws.title}, idx={idx}): {e}")
                    import traceback
                    traceback.print_exc()
                    continue

        wb.close()
        print(f"[画像抽出] 合計: {len(images)}枚")

    except Exception as e:
        print(f"Excelファイル読み込みエラー: {e}")
        import traceback
        traceback.print_exc()
        return []

    return images


def save_images_to_directory(
    file_path: Path,
    output_dir: Path,
    sheet_name: Optional[str] = None,
    prefix: str = "banner"
) -> List[Path]:
    """
    Excelファイルから画像を抽出してディレクトリに保存

    Args:
        file_path: Excelファイルパス
        output_dir: 出力ディレクトリ
        sheet_name: シート名（指定しない場合は全シート）
        prefix: ファイル名プレフィックス

    Returns:
        保存したファイルパスのリスト
    """
    images = extract_images_from_excel(file_path, sheet_name)

    output_dir.mkdir(parents=True, exist_ok=True)
    saved_paths = []

    for img_info in images:
        # ファイル名を生成
        filename = f"{prefix}_{img_info['sheet']}_{img_info['image_index']}.{img_info['format']}"
        output_path = output_dir / filename

        # Base64データをデコードして保存
        image_data = base64.b64decode(img_info['data'])
        output_path.write_bytes(image_data)

        saved_paths.append(output_path)

    return saved_paths


def get_banner_info_table(file_path: Path, sheet_name: Optional[str] = None) -> str:
    """
    バナー画像のメタ情報をMarkdown形式のテーブルで取得（AI考察用）

    Args:
        file_path: Excelファイルパス
        sheet_name: シート名

    Returns:
        Markdown形式のテーブル
    """
    images = extract_images_from_excel(file_path, sheet_name)

    if not images:
        print("[バナー情報テーブル] 画像が見つかりませんでした")
        return ""

    print(f"[バナー情報テーブル] {len(images)}枚の画像を処理中")

    lines = []
    lines.append("## バナー画像情報")
    lines.append("")
    lines.append("| # | 媒体 | サイズ |")
    lines.append("|---|------|------|")

    for idx, img_info in enumerate(images, 1):
        width = img_info['width']
        height = img_info['height']
        media = img_info.get('media_type', '不明')

        lines.append(f"| {idx} | {media} | {width}px × {height}px |")

    lines.append("")
    return "\n".join(lines)


def get_banner_images_markdown(file_path: Path, sheet_name: Optional[str] = None) -> str:
    """
    バナー画像をMarkdown形式で取得（ブラウザ表示用）

    Args:
        file_path: Excelファイルパス
        sheet_name: シート名

    Returns:
        Markdown形式の画像埋め込みテキスト
    """
    images = extract_images_from_excel(file_path, sheet_name)

    if not images:
        print("[バナー画像プレビュー] 画像が見つかりませんでした")
        return ""

    print(f"[バナー画像プレビュー] {len(images)}枚の画像を埋め込み中")

    lines = []
    lines.append("## バナー画像（プレビュー）")
    lines.append("")

    for idx, img_info in enumerate(images, 1):
        media = img_info.get('media_type', '不明')
        # Base64データURLを生成
        data_url = f"data:image/{img_info['format']};base64,{img_info['data']}"

        # データサイズをログ出力
        data_size_kb = len(img_info['data']) / 1024
        print(f"  → バナー#{idx}: {media} / {img_info['format'].upper()} / {data_size_kb:.1f}KB")

        lines.append(f"### バナー #{idx} - {media} ({img_info['sheet']})")
        lines.append(f"**サイズ**: {img_info['width']}x{img_info['height']}px / **形式**: {img_info['format'].upper()}")
        lines.append("")
        lines.append(f"![バナー{idx}]({data_url})")
        lines.append("")

    return "\n".join(lines)


def extract_ad_titles_fallback(file_path: Path, sheet_name: Optional[str] = None) -> str:
    """
    画像抽出が失敗した場合の代替：広告タイトルをテキストで抽出

    Args:
        file_path: Excelファイルパス
        sheet_name: シート名（指定しない場合は全シート）

    Returns:
        Markdown形式の広告タイトル一覧
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets_to_process = [wb[sheet_name]] if sheet_name else wb.worksheets

        lines = []
        lines.append("## 広告クリエイティブ情報（テキスト）")
        lines.append("")
        lines.append("> ℹ️ 画像データを取得できなかったため、広告タイトル・テキスト情報を抽出しました")
        lines.append("")

        found_any = False

        for ws in sheets_to_process:
            # 媒体種別を推定
            media_type = "不明"
            sheet_lower = ws.title.lower()
            if "fb" in sheet_lower or "facebook" in sheet_lower:
                media_type = "Facebook"
            elif "ydn" in sheet_lower or "yahoo" in sheet_lower:
                media_type = "YDN"
            elif "gdn" in sheet_lower or "google" in sheet_lower:
                media_type = "GDN"
            elif "youtube" in sheet_lower or "yt" in sheet_lower:
                media_type = "YouTube"
            elif "デマンド" in ws.title or "demand" in sheet_lower:
                media_type = "デマンド"
            elif "imgad" in sheet_lower:
                media_type = "画像広告"

            # 広告タイトルやクリエイティブ名を探す
            # ヘッダー行を探索（最初の20行以内）
            header_keywords = ["広告名", "クリエイティブ", "タイトル", "広告文", "見出し", "ad name", "creative", "title", "headline"]
            title_col = None
            header_row = None

            for row_idx in range(1, min(21, ws.max_row + 1)):
                for col_idx in range(1, min(11, ws.max_column + 1)):
                    cell_value = ws.cell(row_idx, col_idx).value
                    if cell_value:
                        cell_str = str(cell_value).lower().strip()
                        for keyword in header_keywords:
                            if keyword in cell_str:
                                title_col = col_idx
                                header_row = row_idx
                                break
                    if title_col:
                        break
                if title_col:
                    break

            if title_col and header_row:
                # タイトル列のデータを収集（ヘッダーの次の行から最大50件）
                titles = []
                for row_idx in range(header_row + 1, min(header_row + 51, ws.max_row + 1)):
                    title_value = ws.cell(row_idx, title_col).value
                    if title_value and str(title_value).strip():
                        titles.append(str(title_value).strip())

                if titles:
                    found_any = True
                    lines.append(f"### {media_type} - {ws.title}")
                    lines.append("")
                    for idx, title in enumerate(titles[:10], 1):  # 最大10件表示
                        lines.append(f"{idx}. {title}")
                    if len(titles) > 10:
                        lines.append(f"... 他{len(titles) - 10}件")
                    lines.append("")

        wb.close()

        if not found_any:
            lines.append("広告タイトル情報が見つかりませんでした。")
            lines.append("")

        return "\n".join(lines)

    except Exception as e:
        print(f"広告タイトル抽出エラー: {e}")
        import traceback
        traceback.print_exc()
        return ""
